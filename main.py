import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox
import warnings
import os
import sys
import time

# Отключаем все предупреждения
warnings.filterwarnings('ignore')

# Цвета для заливки
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Светло-зеленый
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Светло-желтый
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Светло-красный

# Стиль выравнивания по центру
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def normalize_name(fio):
    """Нормализует ФИО для сравнения"""
    if pd.isna(fio):
        return ''
    # Приводим к нижнему регистру и удаляем лишние пробелы
    return ' '.join(str(fio).lower().split())


def create_fio_from_columns(row):
    """Создает ФИО из отдельных колонок"""
    parts = []
    if pd.notna(row.get('Фамилия')):
        parts.append(str(row['Фамилия']).strip())
    if pd.notna(row.get('Имя')):
        parts.append(str(row['Имя']).strip())
    if pd.notna(row.get('Отчество')):
        parts.append(str(row['Отчество']).strip())
    return ' '.join(parts)


def is_file_locked(filepath):
    """Проверяет, заблокирован ли файл (открыт в другой программе)"""
    import platform

    if platform.system() == 'Windows':
        try:
            # На Windows пробуем открыть файл в режиме исключительного доступа
            handle = os.open(filepath, os.O_RDWR | os.O_EXCL)
            os.close(handle)
            return False
        except OSError:
            return True
    else:
        # На Linux/Mac используем lsof
        import subprocess
        try:
            subprocess.check_output(['lsof', filepath])
            return True
        except subprocess.CalledProcessError:
            return False


def adjust_column_width(ws):
    """Автоматически настраивает ширину колонок с расширенными настройками"""
    # Ширины для разных типов колонок
    COLUMN_WIDTHS = {
        'процент': 25,  # Процент совпадения
        'источник': 15,  # Источник
        'статус': 35,  # Статус совпадения (увеличен)
        'фио': 45,  # ФИО
        'совпадение': 45,  # Совпадение с порталом
        'default': 20  # Остальные колонки
    }

    # Определяем ширину для каждой колонки
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        column_name = str(ws[f"{column_letter}1"].value).lower() if ws[f"{column_letter}1"].value else ""

        # Определяем тип колонки и устанавливаем ширину
        width = COLUMN_WIDTHS['default']

        if any(keyword in column_name for keyword in ['процент', '%']):
            width = COLUMN_WIDTHS['процент']
        elif 'источник' in column_name:
            width = COLUMN_WIDTHS['источник']
        elif any(keyword in column_name for keyword in ['статус', 'результат', 'проверки']):
            width = COLUMN_WIDTHS['статус']
        elif any(keyword in column_name for keyword in ['фио', 'фамилия', 'имя', 'отчество']):
            width = COLUMN_WIDTHS['фио']
        elif any(keyword in column_name for keyword in ['совпадение', 'match', 'соответствие']):
            width = COLUMN_WIDTHS['совпадение']

        ws.column_dimensions[column_letter].width = width

        # Выравниваем заголовок по центру
        ws[f"{column_letter}1"].alignment = CENTER_ALIGNMENT

        # Для колонки статуса также выравниваем содержимое по центру
        if 'статус' in column_name:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{column_letter}{row}"]
                cell.alignment = CENTER_ALIGNMENT


def save_with_formatting(filepath, df):
    """Сохраняет DataFrame с форматированием"""
    try:
        # Сначала сохраняем без форматирования
        df.to_excel(filepath, index=False)

        # Теперь применяем форматирование
        wb = load_workbook(filepath)
        ws = wb.active

        # Настраиваем ширину колонок
        adjust_column_width(ws)

        # Находим индекс колонки с процентами
        header = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        percent_col_idx = None

        for idx, col_name in enumerate(header, 1):
            if 'процент' in str(col_name).lower():
                percent_col_idx = idx
                break

        # Применяем выравнивание по центру для всей колонки процентов
        if percent_col_idx:
            col_letter = get_column_letter(percent_col_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.alignment = CENTER_ALIGNMENT

        # Применяем цветовую разметку
        apply_coloring_to_worksheet(ws)

        # Сохраняем изменения
        wb.save(filepath)
        print(f"Файл с форматированием сохранен: {filepath}")
        return True

    except Exception as e:
        print(f"Ошибка при сохранении с форматированием: {e}")
        # Пробуем сохранить без форматирования
        try:
            df.to_excel(filepath, index=False)
            print(f"Файл сохранен без форматирования: {filepath}")
            return True
        except Exception as e2:
            print(f"Ошибка при сохранении без форматирования: {e2}")
            return False


def process_excel_file(input_file, threshold=85):
    """
    Основная функция обработки Excel файла
    threshold - порог частичного совпадения (85 по умолчанию)
    """
    # Читаем Excel файл
    print(f"Чтение файла: {input_file}")

    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        raise ValueError(f"Ошибка при чтении файла: {e}")

    # Проверяем наличие необходимых колонок
    required_columns = ['источник']

    # Проверяем наличие колонок ФИО
    fio_columns_check = []

    # Сначала проверим, есть ли отдельные колонки ФИО
    if all(col in df.columns for col in ['Фамилия', 'Имя', 'Отчество']):
        fio_columns_check = ['Фамилия', 'Имя', 'Отчество']
        print("Найдены отдельные колонки ФИО: Фамилия, Имя, Отчество")
    else:
        # Ищем колонку с полным ФИО
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in ['фио', 'фам', 'фамилия', 'полное']):
                fio_columns_check = [col]
                print(f"Найдена колонка с ФИО: {col}")
                break

    if not fio_columns_check:
        raise ValueError("Не найдены колонки с ФИО. Нужны либо 'Фамилия', 'Имя', 'Отчество', либо колонка с полным ФИО")

    # Создаем временную колонку с полным ФИО
    if len(fio_columns_check) == 3:
        df['_temp_ФИО'] = df.apply(create_fio_from_columns, axis=1)
        fio_column = '_temp_ФИО'
    else:
        fio_column = fio_columns_check[0]
        df['_temp_ФИО'] = df[fio_column].astype(str)

    print(f"Используем ФИО из колонки: {fio_column}")

    # Нормализуем источник
    df['источник_норм'] = df['источник'].astype(str).str.lower().str.strip()

    # Разделяем данные по источникам
    zups = df[df['источник_норм'].str.contains('зуп', na=False)].copy()
    portal = df[df['источник_норм'].str.contains('портал', na=False)].copy()

    print(f"Найдено записей в ЗУП: {len(zups)} (основной источник)")
    print(f"Найдено записей в Портал: {len(portal)} (сравниваем с ЗУП)")

    if len(zups) == 0:
        raise ValueError("Не найдено записей с источником 'ЗУП'")
    if len(portal) == 0:
        raise ValueError("Не найдено записей с источником 'портал'")

    # Создаем словарь нормализованных ФИО из ПОРТАЛА
    portal_fios_dict = {}
    for idx, row in portal.iterrows():
        fio = row['_temp_ФИО']
        if pd.notna(fio) and str(fio).strip():
            normalized = normalize_name(fio)
            if normalized:
                portal_fios_dict[normalized] = {
                    'original_fio': fio,
                    'row_idx': idx
                }

    print(f"Создан словарь из портала: {len(portal_fios_dict)} уникальных ФИО")

    # Создаем список для результатов
    results = []

    # Проверяем каждую запись из ЗУП
    for idx, row in zups.iterrows():
        zup_fio = row['_temp_ФИО']
        original_zup_data = row.to_dict()

        if pd.isna(zup_fio) or not str(zup_fio).strip():
            results.append({
                'row_idx': idx,
                'источник': 'ЗУП',
                'фио_в_зуп': zup_fio if pd.notna(zup_fio) else '',
                'совпадение_с_порталом': '',
                'процент_совпадения': 0,
                'статус_совпадения': 'Пустое ФИО в ЗУП'
            })
            continue

        normalized_zup = normalize_name(zup_fio)

        # Ищем точное совпадение
        exact_match = None
        if normalized_zup in portal_fios_dict:
            exact_match = portal_fios_dict[normalized_zup]
            match_fio = exact_match['original_fio']
            match_score = 100
            status = 'Полное совпадение'

            results.append({
                'row_idx': idx,
                'источник': 'ЗУП',
                'фио_в_зуп': zup_fio,
                'совпадение_с_порталом': match_fio,
                'процент_совпадения': match_score,
                'статус_совпадения': status
            })

            # Удаляем из словаря
            del portal_fios_dict[normalized_zup]
            continue

        # Ищем лучшее нечеткое совпадение
        best_match = ''
        best_score = 0
        best_key = None

        for portal_key, portal_data in portal_fios_dict.items():
            score = fuzz.token_sort_ratio(normalized_zup, portal_key)
            if score > best_score:
                best_score = score
                best_match = portal_data['original_fio']
                best_key = portal_key

        if best_score >= threshold:
            status = 'Частичное совпадение'
        else:
            status = 'Совпадений не найдено'
            best_match = ''

        results.append({
            'row_idx': idx,
            'источник': 'ЗУП',
            'фио_в_зуп': zup_fio,
            'совпадение_с_порталом': best_match,
            'процент_совпадения': best_score,
            'статус_совпадения': status
        })

        # Удаляем найденное совпадение из словаря
        if best_key and best_score >= threshold:
            del portal_fios_dict[best_key]

    # Добавляем записи из портала, которые не нашли совпадений в ЗУП
    for portal_key, portal_data in portal_fios_dict.items():
        results.append({
            'row_idx': portal_data['row_idx'],
            'источник': 'портал',
            'фио_в_зуп': '',
            'совпадение_с_порталом': portal_data['original_fio'],
            'процент_совпадения': 0,
            'статус_совпадения': 'Нет в ЗУП'
        })

    # Создаем DataFrame с результатами
    results_df = pd.DataFrame(results)

    # Сортируем: сначала ЗУП, потом портал
    results_df['sort_key'] = results_df['источник'].apply(lambda x: 0 if x == 'ЗУП' else 1)
    results_df = results_df.sort_values(['sort_key', 'row_idx']).drop('sort_key', axis=1)

    # Добавляем оригинальные данные
    output_data = []

    for _, result_row in results_df.iterrows():
        original_idx = result_row['row_idx']
        original_row = df.loc[original_idx]

        # Создаем строку для вывода
        output_row = {
            'источник': original_row['источник'],
            'статус_совпадения': result_row['статус_совпадения'],
            'совпадение_с_порталом': result_row['совпадение_с_порталом'],
            'процент_совпадения': result_row['процент_совпадения'],
            'фио_в_зуп': result_row['фио_в_зуп']
        }

        # Добавляем остальные колонки
        for col in df.columns:
            col_lower = str(col).lower()
            if (col_lower in ['источник', 'фамилия', 'имя', 'отчество', '_temp_фио', 'источник_норм'] or
                    'unnamed' in col_lower):
                continue
            if col not in output_row:
                output_row[col] = original_row[col]

        output_data.append(output_row)

    # Создаем итоговый DataFrame
    final_df = pd.DataFrame(output_data)

    # Генерируем имя выходного файла
    base_name = os.path.splitext(input_file)[0]
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_file = f"{base_name}_результат_{timestamp}.xlsx"

    # Сохраняем результаты с форматированием
    print(f"Сохранение результатов в: {output_file}")

    if is_file_locked(output_file) and os.path.exists(output_file):
        # Если файл существует и заблокирован, создаем новый с другим именем
        output_file = f"{base_name}_результат_{timestamp}_new.xlsx"
        print(f"Создаю новый файл: {output_file}")

    success = save_with_formatting(output_file, final_df)

    if not success:
        # Пробуем еще раз с другим именем
        output_file = f"{base_name}_результат_{timestamp}_final.xlsx"
        print(f"Пробую сохранить как: {output_file}")
        success = save_with_formatting(output_file, final_df)

        if not success:
            raise Exception("Не удалось сохранить файл после нескольких попыток")

    return output_file, final_df


def apply_coloring_to_worksheet(ws):
    """
    Применяем цветовую разметку к рабочему листу
    """
    try:
        print(f"Начинаю раскраску рабочего листа")

        # Находим индексы колонок по заголовкам
        header = []
        for cell in ws[1]:  # Первая строка
            header.append(str(cell.value).strip() if cell.value else '')

        print(f"Заголовки колонок: {header}")

        # Ищем индексы колонок
        status_col_idx = None
        source_col_idx = None

        for idx, col_name in enumerate(header, 1):
            col_name_str = str(col_name).lower()
            if 'статус' in col_name_str:
                status_col_idx = idx
                print(f"Найдена колонка статуса: {col_name} (индекс {idx})")
            elif 'источник' in col_name_str:
                source_col_idx = idx
                print(f"Найдена колонка источника: {col_name} (индекс {idx})")

        if not status_col_idx:
            print("Предупреждение: колонка статуса не найдена, ищем по другим ключевым словам")
            for idx, col_name in enumerate(header, 1):
                if any(word in str(col_name).lower() for word in ['совпадения', 'результат', 'проверки']):
                    status_col_idx = idx
                    print(f"Найдена колонка по альтернативному ключу: {col_name}")
                    break

        if not source_col_idx:
            print("Ошибка: не найдена колонка 'источник'")
            return

        if not status_col_idx:
            print("Ошибка: не найдена колонка статуса")
            return

        print(f"Используем колонку источника: индекс {source_col_idx}")
        print(f"Используем колонку статуса: индекс {status_col_idx}")

        # Применяем цветовую разметку
        colored_count = 0
        total_rows = ws.max_row

        for row_num in range(2, total_rows + 1):
            source_cell = ws.cell(row=row_num, column=source_col_idx)
            status_cell = ws.cell(row=row_num, column=status_col_idx)

            source_value = str(source_cell.value).lower() if source_cell.value else ''
            status_value = str(status_cell.value) if status_cell.value else ''

            if 'зуп' in source_value:
                if status_value == 'Полное совпадение':
                    fill_color = GREEN_FILL
                elif status_value == 'Частичное совпадение':
                    fill_color = YELLOW_FILL
                elif status_value in ['Совпадений не найдено', 'Пустое ФИО в ЗУП']:
                    fill_color = RED_FILL
                else:
                    fill_color = None

                if fill_color:
                    # Закрашиваем всю строку
                    for col_num in range(1, len(header) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = fill_color
                    colored_count += 1
            elif 'портал' in source_value:
                # Для портала оставляем без заливки
                pass

        print(f"Раскрашено строк ЗУП: {colored_count}")
        print(f"Всего строк в файле: {total_rows - 1}")

        # Добавляем автофильтр к заголовкам
        if total_rows > 1:
            ws.auto_filter.ref = ws.dimensions

        print("Раскраска и форматирование завершены успешно!")

    except Exception as e:
        print(f"Ошибка при раскраске: {e}")
        import traceback
        traceback.print_exc()


def select_file():
    """Функция для выбора файла через диалоговое окно"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    file_path = filedialog.askopenfilename(
        title="Выберите Excel файл для обработки",
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )

    return file_path


def create_settings_window():
    """Создает окно настроек для выбора порога совпадения"""

    def on_submit():
        try:
            threshold = int(threshold_var.get())
            if 0 <= threshold <= 100:
                window.threshold = threshold
                window.destroy()
            else:
                messagebox.showerror("Ошибка", "Порог должен быть от 0 до 100")
        except ValueError:
            messagebox.showerror("Ошибка", "Введите число от 0 до 100")

    window = tk.Tk()
    window.title("Настройки обработки")
    window.geometry("400x200")

    # Центрируем окно
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

    tk.Label(window, text="Порог частичного совпадения", font=("Arial", 12)).pack(pady=10)
    tk.Label(window, text="(от 0 до 100, по умолчанию 85)").pack()

    threshold_var = tk.StringVar(value="85")
    entry = tk.Entry(window, textvariable=threshold_var, font=("Arial", 12), width=10)
    entry.pack(pady=10)

    tk.Button(window, text="Начать обработку", command=on_submit, width=15, height=2).pack(pady=20)

    window.threshold = 85
    window.mainloop()

    return getattr(window, 'threshold', 85)


def show_results_window(output_file, df):
    """Показывает окно с результатами обработки"""
    root = tk.Tk()
    root.title("Результаты обработки")
    root.geometry("620x720")

    # Центрируем окно
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    # Главный фрейм
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Заголовок
    tk.Label(main_frame, text="Обработка завершена успешно!",
             font=("Arial", 14, "bold")).pack(pady=5)

    # Информация о файле
    file_frame = tk.LabelFrame(main_frame, text="Сохраненный файл", padx=10, pady=10)
    file_frame.pack(fill=tk.X, pady=5)

    file_label = tk.Label(file_frame, text=output_file, wraplength=550, justify=tk.LEFT)
    file_label.pack()

    # Кнопка для открытия папки
    def open_folder():
        folder = os.path.dirname(output_file)
        if os.path.exists(folder):
            os.startfile(folder) if sys.platform == 'win32' else os.system(f'open "{folder}"')

    tk.Button(file_frame, text="Открыть папку с файлом",
              command=open_folder).pack(pady=5)

    # Кнопка для открытия файла
    def open_file():
        if os.path.exists(output_file):
            os.startfile(output_file) if sys.platform == 'win32' else os.system(f'open "{output_file}"')

    tk.Button(file_frame, text="Открыть файл с результатами",
              command=open_file).pack(pady=5)

    # Статистика по ЗУП
    stats_frame = tk.LabelFrame(main_frame, text="Статистика для ЗУП", padx=10, pady=10)
    stats_frame.pack(fill=tk.X, pady=5)

    if 'статус_совпадения' in df.columns:
        zup_data = df[df['источник'].astype(str).str.lower().str.contains('зуп', na=False)]

        if len(zup_data) > 0:
            stats = zup_data['статус_совпадения'].value_counts()

            for result, count in stats.items():
                frame = tk.Frame(stats_frame)
                frame.pack(fill=tk.X, pady=2)

                # Цветной индикатор
                if result == 'Полное совпадение':
                    color = 'green'
                elif result == 'Частичное совпадение':
                    color = 'orange'
                elif result == 'Совпадений не найдено':
                    color = 'red'
                elif result == 'Пустое ФИО в ЗУП':
                    color = 'gray'
                else:
                    color = 'black'

                tk.Label(frame, text="●", fg=color, font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
                tk.Label(frame, text=f"{result}: {count} записей",
                         font=("Arial", 10)).pack(side=tk.LEFT)

    # Статистика по порталу
    if 'статус_совпадения' in df.columns:
        portal_data = df[df['источник'].astype(str).str.lower().str.contains('портал', na=False)]
        if len(portal_data) > 0:
            portal_not_found = portal_data[portal_data['статус_совпадения'] == 'Нет в ЗУП']
            if len(portal_not_found) > 0:
                portal_frame = tk.LabelFrame(main_frame, text="Записи портала без совпадений", padx=10, pady=10)
                portal_frame.pack(fill=tk.X, pady=5)

                tk.Label(portal_frame, text=f"Записей портала без совпадений в ЗУП: {len(portal_not_found)}",
                         font=("Arial", 10)).pack()

    # Инструкция по цветам
    instr_frame = tk.LabelFrame(main_frame, text="Инструкция по цветам", padx=10, pady=10)
    instr_frame.pack(fill=tk.X, pady=5)

    instructions = [
        ("ЗЕЛЕНЫЙ", "Записи ЗУП с полным совпадением в портале"),
        ("ЖЕЛТЫЙ", "Записи ЗУП с частичным совпадением в портале"),
        ("КРАСНЫЙ", "Записи ЗУП без совпадений в портале"),
        ("БЕЛЫЙ", "Записи портала (не раскрашиваются)")
    ]

    for color, text in instructions:
        frame = tk.Frame(instr_frame)
        frame.pack(fill=tk.X, pady=2)
        tk.Label(frame, text=color, font=("Arial", 10, "bold"),
                 width=35, anchor=tk.W).pack(side=tk.LEFT)
        tk.Label(frame, text=text).pack(side=tk.LEFT)

    # Информация о форматировании
    format_frame = tk.LabelFrame(main_frame, text="Форматирование файла", padx=10, pady=10)
    format_frame.pack(fill=tk.X, pady=5)

    tk.Label(format_frame, text="• Ширина колонок автоматически настроена",
             font=("Arial", 9)).pack(anchor=tk.W)
    tk.Label(format_frame, text="• Процент совпадения выровнен по центру",
             font=("Arial", 9)).pack(anchor=tk.W)
    tk.Label(format_frame, text="• Добавлен автофильтр к заголовкам",
             font=("Arial", 9)).pack(anchor=tk.W)

    # Кнопка закрытия
    tk.Button(main_frame, text="Закрыть", command=root.destroy,
              width=20, height=2).pack(pady=10)

    root.mainloop()


def main():
    print("=" * 50)
    print("Программа для сравнения: ЗУП (основной) vs Портал")
    print("Раскрашиваются только записи ЗУП")
    print("=" * 50)

    try:
        # Показываем окно настроек
        threshold = create_settings_window()
        print(f"Установлен порог совпадения: {threshold}%")

        # Выбираем файл
        input_file = select_file()

        if not input_file:
            print("Файл не выбран. Программа завершена.")
            return

        print(f"Выбран файл: {input_file}")

        # Запуск обработки
        output_file, df = process_excel_file(input_file, threshold)

        # Показываем результаты в графическом окне
        show_results_window(output_file, df)

    except ValueError as e:
        messagebox.showerror("Ошибка", str(e))
        print(f"Ошибка: {e}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка при обработке файла:\n{str(e)}")
        print(f"Произошла ошибка: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()